import numpy as np
import os
import xlrd
import openpyxl
from closed_class import closed_classes
import hittite_pos_tagger


def make_predict_set():
    fname_list = os.listdir('../predict/') # массив имён файлов с обучающими материалами в .xlsx

    closed = closed_classes()
    feature_set = [] # матрица признаков
    speech_class = {i.split()[1]: i.split()[0] for i in open('../class.txt').read().split('\n') if i != ''}
    # считываем части речи с их порядковым номером в словарь
    class_speech = {i.split()[0]: i.split()[1] for i in open('../class.txt').read().split('\n') if i != ''}
    features = open('../features.txt').read().split()
    # считываем признаки
    fw = open('prediction_words.txt', 'w', encoding='utf-8') # файл со словами, для которых будет предсказываться часть речи

    models = [m for m in hittite_pos_tagger.pos_tagger()]

    for fname in fname_list: # для каждого файла

        print('processing ' + fname + '\n')

        rb = xlrd.open_workbook('../predict/' + fname) # читаем дерево
        sheet = rb.sheet_by_index(0) # идём в лист 'Syntax'
        sentences = []

        for rownum in range(2, sheet.nrows): # обходим значения каждого ряда с третьего, потому что первые два -- названия колонок
            row = sheet.row_values(rownum)
            sentence_data = [row[12], row[6], row[9], row[10]] # массив с предложениями и метаданными
            sentences.append(sentence_data)

        wb = openpyxl.load_workbook(filename='../predict/' + fname)  # читаем дерево
        rownum = 3  # начинаем с третьего ряда, ибо первые два -- названия колонок

        sheet = wb['Word Forms'] # идём в лист 'Word Forms'

        for sentence in sentences:
            words = sentence[0].split()
            for indx, w in enumerate(words): # делим по пробелам и равно, сохраняя равно
                if '=' in w[1:]:
                    i = w[1:].find('=')
                    words[indx] = w[:i + 1]
                    words.insert(indx + 1, w[i + 1:])

            for ind, word in enumerate(words):
                sheet.cell(column=1, row=rownum, value=sentence[1]) # добавляем Text index
                sheet.cell(column=2, row=rownum, value=sentence[2]) # Line
                sheet.cell(column=3, row=rownum, value=sentence[3]) # Clause index
                sheet.cell(column=4, row=rownum, value='ZZ' + str(ind+1)) # Word Form index
                sheet.cell(column=6, row=rownum, value=word.replace('=', '- -')) # narrow transliteration
                sheet.cell(column=11, row=rownum, value=ind+1) # WordOrder

                if word.replace('…', '...') in closed.keys():
                    sheet.cell(column=7, row=rownum, value=closed[word.replace('…', '...')])  # Glosses
                    sheet.cell(column=9, row=rownum, value=closed[word.replace('…', '...')].split('.')[::-1][0].upper())  # PoS
                else:
                    fw.write(word)
                    example = []
                    for feat in features: # добавляем его признаки
                        if word.strip('!: ').lower().endswith(feat):
                            example.append(1.0) # есть признак
                        else:
                            example.append(0.0) # нет признака
                    example.append(float(ind+1)) # добавляем номер в предложении
                    feature_set.append(feature_set)
                    x = np.array(example).reshape(1,-1)

                    for m in models:
                        tag = m.predict(x)
                        fw.write(' ' + class_speech[np.asscalar(tag)])

                    tag = models[2].predict(x)
                    sheet.cell(column=9, row=rownum, value=class_speech[np.asscalar(tag)])

                    fw.write('\n')
                rownum += 1

        wb.save(filename='../predict/' + fname)
    fw.close()


def main():
    make_predict_set()


if __name__ == '__main__':
    main()